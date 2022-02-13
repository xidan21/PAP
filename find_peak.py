
import re
import os
import sys
import operator
from scipy import stats
import numpy as np
import openpyxl
filename = sys.argv[1]
print sys.argv[2]
gap = int(sys.argv[2])


wb = openpyxl.load_workbook(filename)

z = 0


print sys.argv[1]

print "SHEET name"+"\t"+"Condition number"+"\t"+"Peak Value"+"\t"+"50% time elapse"+"\t"+"75% time elapse"+"\t"+"Amount of peaks"+ "\t" + "Average of up velocity" + "\t" +"Average of decay velocity" + "\t" + "Average of peak baseline"
for sheet in wb.worksheets:

	peak_value_array = []
        peak_50_array = []
        peak_75_array = []
        peak_amount_array = []
        peak_ave_up_velocity = []
        peak_ave_down_velocity = []
	peak_baseline = []
	
	SHEET_NAME = wb.worksheets[z]

        for b in xrange(3, (sheet.max_column)+1):
		times = []
                x = []
                for a in xrange(2, (sheet.max_row)+1):

                        times.append(sheet.cell(row=a, column=2).value)
                        x.append(sheet.cell(row=a, column=b).value)


		times[0] = 0

                first_point = times[0]

                last_point = times[-1]
                
                total_frequency = (float(last_point) - float(first_point)) / float(len(times))

		ORIGINAL_X = x
		###################### normalize the values #######################

		MIN_X = min(x)

		distance_array = []

		for i in xrange(len(x)-1):

	#		print x[i],x[i+1]

			distance_array.append(abs(x[i]-x[i+1]))

	#		print (abs(x[i] - x[i+1]))

		AVE_DISTANCE =  float(sum(distance_array)) / float(len(distance_array))

		###################### identify the start and end positions of curves  #######################


		peak_shape = []


		peak_start_end_value = []

		bla_count = 0
		count = 0

		for i in xrange(0, len(x)-gap-gap):

			slope_count = 0

			first_group = []
	
			second_group = []

			for j in xrange(0,gap):
			
				first_group.append(x[i+j])

			for k in xrange(gap, gap+gap):

				second_group.append(x[i+k])
		
			if operator.gt (second_group[0], first_group[-1]):

				DISTANCE_first = []
				
				for l in xrange(len(first_group)-1):
			
					DISTANCE_first.append(float(first_group[l+1]) - float(first_group[l]))
                               
				DISTANCE_second = []

                                for l in xrange(len(second_group)-1):

                                        DISTANCE_second.append(float(second_group[l+1]) - float(second_group[l]))


				for m in xrange(len(DISTANCE_second)):
				
					if operator.gt(DISTANCE_second[m], 0):			
						slope_count += 1

				#if operator.eq(gap, slope_count+1):
				if operator.eq(gap, slope_count+1) and operator.gt(float(sum(DISTANCE_second))/float(len(DISTANCE_second)), float(AVE_DISTANCE)):
					count += 1

					peak_shape.append(i+gap-2) ## end_value
					peak_shape.append(i+gap-1) # start_value
						

                ##################### remove the false peak #####################

                # 2562, 3276, 3277, 3277, 3278, 3516,

	        to_delete = []

       		for i in xrange(len(peak_shape)-2):

                	if operator.eq(peak_shape[i], peak_shape[i+1]):

                        	to_delete.append(i+1)
                        	to_delete.append(i+2)

        	for offset, index in enumerate(to_delete):

                	index -= offset
        
                	del peak_shape[index]

	#	print peak_shape

               # to delete the small fake peak

                PEAK_DISTANCE = []

                for i in xrange(1, len(peak_shape)-1, 2):

                        PEAK_DISTANCE.append(float(peak_shape[i+1]) - float(peak_shape[i]))


           #     print PEAK_DISTANCE

		AVE_PEAK_DISTANCE = float(sum(PEAK_DISTANCE)) / float(len(PEAK_DISTANCE))

		to_delete_2 = []
		for i in xrange (len(PEAK_DISTANCE)):


			if operator.gt (float(AVE_PEAK_DISTANCE), float(PEAK_DISTANCE[i])*1.5):

				to_delete_2.append(i*2)
				to_delete_2.append(i*2+1)


                for offset, index in enumerate(to_delete_2):

                	index -= offset
 
                        del peak_shape[index]
 

       # 	print peak_shape
 #               print len(peak_shape)
		##################### normalize the curve inside #######################


		for i in xrange(1, len(peak_shape)-1, 2):


			temp_time = []
			temp_curve = []

			for j in xrange(int(peak_shape[i]), int(peak_shape[i+1])+1):
		#		print j
				temp_curve.append(x[j])
				temp_time.append(j)

		#	print temp_time




			slope_length = (peak_shape[i+1] - peak_shape[i]) * total_frequency

			slope_height = temp_curve[0] - temp_curve[-1]

			slope_ratio = abs(float(slope_height)) / float(slope_length)

#			print peak_shape[i+1]
#			print peak_shape[i]

#			print total_frequency

#			print "slope_ratio", str(slope_ratio)

#			print "slope_length", str(slope_length)

#			print temp_curve[0]
#			print temp_curve[-1]

#			print "slope_height", str(slope_height)

 #                       print "########## before ###########"
  #                      print x[peak_shape[i]]
   #                     print x[peak_shape[i+1]]


			if operator.gt (slope_height, 0):

				for k in xrange(len(temp_curve)):
					
					# the distance of each point to the end curve: (peak_shape[k]) - temp_time[-1]) * total_frequency 
				
					# the height of each point minus in order to normalize ( ((peak_shape[k]) - temp_curve[-1]) * total_frequency) * slope_ratio)
				
				#	print (peak_shape[i])+k
                              #          print x[(peak_shape[i])+k]

                               #         print "temp_curve[k]", temp_curve[k]

                                #        print ((((peak_shape[i])+k - temp_time[-1]) * total_frequency) * slope_ratio)

					x[(peak_shape[i])+k] = temp_curve[k] - ((abs((peak_shape[i])+k - temp_time[-1]) * total_frequency) * slope_ratio)
				#	print x[(peak_shape[i])+k]

			elif operator.lt (slope_height, 0):

				for k in xrange(len(temp_curve)):
					
					# the distance of each point to the start curve: (peak_shape[k]) - temp_time[0]) * total_frequency 

					# the height of each point minus in order to normalize (((peak_shape[k]) - temp_curve[0]) * total_frequency) * slope_ratio)

				#       print (peak_shape[i])+k

			#		print x[(peak_shape[i])+k]

			#		print "temp_curve[k]", temp_curve[k]

			#		print ((((peak_shape[i])+k - temp_time[0]) * total_frequency) * slope_ratio)

					x[(peak_shape[i])+k] = temp_curve[k] - ((((peak_shape[i])+k - temp_time[0]) * total_frequency) * slope_ratio)

					
			#		print x[(peak_shape[i])+k]

#			print "########## after ###########"	
#			print x[peak_shape[i]] 
#			print x[peak_shape[i+1]]

#			print "############################"
		##################### normalize the between curves #######################

		MIN_x = min(x)

		MIN_pos = x.index(min(x))

		#print MIN_pos
		#print MIN_x

		#print peak_shape

		for i in xrange(1, len(peak_shape)-1,2):


			global_slope_height = x[peak_shape[i+1]] - MIN_x

			for j in xrange(int(peak_shape[i]), int(peak_shape[i+1])+1):


				if operator.le(j, len(x)):

					x[j] = x[j] - global_slope_height
	#	for i in xrange(1, len(peak_shape)-1, 2):


	#		print "########## after global ###########"
	#		print x[peak_shape[i]]
	#		print x[peak_shape[i+1]]

	#		print "############################"



		##################### set up the windows ######################


		win_start = int(sys.argv[3]) # second: 30

		win_end = int(sys.argv[4]) # second: 120


		windows_time = []

		for i in xrange(len(times)):

			try:
				#print times[i]
				if operator.gt( float(times[i]), float(win_start)) and operator.lt( float(times[i]), float(win_end)):

					windows_time.append(i)

				#	print times[i]

			except IOError as e:

				print 'Oh dear.'

		#print windows_time

		win_start_position =  min(windows_time)
		win_end_position = max(windows_time)
#		print win_start
#		print win_end
#		print ">>>>>>>>>>>>>>>>>>>",win_start_position
#		print ">>>>>>>>>>>>>>>>>>>",win_end_position

		win_peak_shape = []
		total_up_stroke = []
		total_down_stroke = []			

		for i in xrange(0, len(peak_shape), 2):

		#       print peak_shape[i]
#			print win_start_position,">>>>>>>> start >>>>>>>>>"
#			print win_end_position, ">>>>>>> end >>>>>>>"
			if operator.gt (int(peak_shape[i]), win_start_position) and operator.lt (int(peak_shape[i]), win_end_position):

				win_peak_shape.append(peak_shape[i])
				win_peak_shape.append(peak_shape[i+1])

	#	print peak_shape
		
#		print win_peak_shape

		##################### detect the peak value in the windows ######################

		peak_value = []

		MIN_final_x = min(x)

		count = 0
#		print "SHEET name"+"\t"+"Condition number"+"\t"+"Peak Value"+"\t"+"50% time elapse"+"\t"+"75% time elapse"+"\t"+"Amount of peaks"

		array_final_peak = []
		array_final_50 = []
		array_final_90 = []

#		print win_peak_shape

		for i in xrange(1, len(win_peak_shape)-1, 2): ################################################ be careful !!!

			temp_array = []

			count += 1
			bla_count += 1
			for j in xrange(int(win_peak_shape[i]), int(win_peak_shape[i+1])): # one peak in the window

				temp_array.append(x[j])

			#print "PEAK %d" %(count)

			PEAK = max(temp_array)
			START = temp_array[0]
		
			peak_baseline.append(START)		

			END = temp_array[-1]

			peak_baseline.append(END)		
			
			DECAY_75 = (PEAK - END) * 0.75		
			#print DECAY_75

			#print win_peak_shape[i] 

			TIME_START = times[win_peak_shape[i] + temp_array.index(START)]
			TIME_PEAK = times[win_peak_shape[i] + temp_array.index(PEAK)]	
			TIME_END = times[win_peak_shape[i] + temp_array.index(END)]
			TIME_END_75 = (TIME_END - TIME_PEAK) * 0.75
			
			#print TIME_END_75
			#print win_peak_shape[i] + temp_array.index(START), TIME_START
			
			if operator.eq(TIME_PEAK - TIME_START, 0):

				print "the peak is weird!!!"

			else:
				up_velocity = float (PEAK - START) / float (TIME_PEAK - TIME_START)
			
				total_up_stroke.append(up_velocity)

				#print "up velocity is %f" %(up_velocity)

				if operator.eq (TIME_END_75, 0):

					print "the peak is super weird!!!"

				else:
					decay_velocity = float (DECAY_75) / float (TIME_END_75)

					total_down_stroke.append(decay_velocity)

					#print "75% of decay velocity is ", decay_velocity


					up_stroke = []

					for i in xrange(0, temp_array.index(PEAK)):

						up_stroke.append(temp_array[i])

#					print up_stroke

			
					down_stroke = []

					for i in xrange(temp_array.index(PEAK), len(temp_array)):

						down_stroke.append(temp_array[i])

			

#					print down_stroke

		 #      	 	print ">>>>>>>>>>>>>>>>>>>>>>>> the peak %d >>>>>>>>>>>>>>>>>:" %(count)

					array_final_peak.append(float(max(temp_array)) - float(MIN_final_x))

					peak_value_position = temp_array.index(max(temp_array))


					decay_temp_array = temp_array[peak_value_position:]

					decay_length_value = max(temp_array) - temp_array[-1]


					percent_50_value = (max(temp_array) - temp_array[-1]) * 0.5 + temp_array[-1]


					minus_percent_50_index = []

					for i in xrange(len(decay_temp_array)):

						minus_percent_50_index.append(abs(decay_temp_array[i] - percent_50_value))


					array_final_50.append((float(minus_percent_50_index.index(min(minus_percent_50_index))) - 0)*1000 * total_frequency)

			


					percent_90_value = (max(temp_array) - temp_array[-1]) * (1-0.75) + temp_array[-1]


					minus_percent_90_index = []

					for i in xrange(len(decay_temp_array)):

						minus_percent_90_index.append(abs(decay_temp_array[i] - percent_90_value))


		#			print float(minus_percent_90_index.index(min(minus_percent_90_index)))

					array_final_90.append((float(minus_percent_90_index.index(min(minus_percent_90_index))) - 0) * 1000* total_frequency)
		

#				print float(minus_percent_50_index.index(min(minus_percent_50_index)))
#				print total_frequency
#				print (array_final_50)
	#			print (array_final_90)
		#		print SHEET_NAME

				AVE_UP_VELOCITY = sum(total_up_stroke) / len(total_up_stroke)
				AVE_DOWN_VELOCITY = sum(total_down_stroke) / len(total_down_stroke)

				PEAK_BASELINE = (START + END) / 2

				PEAK_VALUE = float(sum(array_final_peak))/float(len(array_final_peak))

		print str(SHEET_NAME)+"\t"+str(b-2)+"\t"+str(PEAK_VALUE)+"\t"+str(float(sum(array_final_50))/float(len(array_final_50))) + "\t" + str(float(sum(array_final_90))/float(len(array_final_90))) + "\t"+ str(bla_count) + "\t" + str(AVE_UP_VELOCITY) + "\t" + str(AVE_DOWN_VELOCITY) + "\t" + str(PEAK_BASELINE)

		
		
		print "Calcium value is equal to " + str(400*PEAK_VALUE / ((400/PEAK_BASELINE + 1) - PEAK_VALUE))

		peak_value_array.append(float(sum(array_final_peak))/float(len(array_final_peak)))
		peak_50_array.append(float(sum(array_final_50))/float(len(array_final_50)))
		peak_75_array.append(float(sum(array_final_90))/float(len(array_final_90)))
		peak_amount_array.append(bla_count)
		peak_ave_up_velocity.append(AVE_UP_VELOCITY)
		peak_ave_down_velocity.append(AVE_DOWN_VELOCITY)
	
	

	if operator.eq(len(peak_value_array),0):

		print "There is no peak detected !!!"
	
	else:

		ave_peak_baseline = sum(peak_baseline) / float(len(peak_baseline))
		ave_peak_value = sum(peak_value_array) / float(len(peak_value_array))
		ave_peak_50 = sum(peak_50_array) / float(len(peak_50_array))
		ave_peak_75 = sum(peak_75_array) / float(len(peak_75_array))
		ave_peak_amount = sum(peak_amount_array) / float(len(peak_amount_array))
		ave_peak_ave_up_velocity = sum(peak_ave_up_velocity) / float(len(peak_ave_up_velocity))
		ave_peak_ave_down_velocity = sum(peak_ave_down_velocity) / float(len(peak_ave_down_velocity))

		print str(SHEET_NAME)+"\t"+"average row "+"\t"+str(ave_peak_value)+"\t"+str(ave_peak_50) + "\t" + str(ave_peak_75) + "\t"+ str(ave_peak_amount) + "\t" + str(ave_peak_ave_up_velocity) + "\t" + str(ave_peak_ave_down_velocity) + "\t" + str(ave_peak_baseline)

		print "Calcium value is equal to " + str(400*ave_peak_value / ((400/ave_peak_baseline + 1) - ave_peak_value))


	bla_count = 0
	z += 1


